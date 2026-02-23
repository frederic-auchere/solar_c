import os
import shutil
import numpy as np
import matplotlib

matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
import pandas as pd
from optical.utils import fit_circle_to_points
from optics.zygo import ZygoData

IGNORED_FOLDERS = {"shutter"}


# ================================================================
# ===================== CONVERSION STEP ==========================
# ================================================================

def convert_qpsix_to_npy(input_folder, output_folder):
    """
    Converts all .qpsix files in input_folder to .npy files in output_folder.
    Skips files that have already been converted.
    """
    os.makedirs(output_folder, exist_ok=True)

    any_processed = False

    for dirpath, dirnames, filenames in os.walk(input_folder):
        dirnames[:] = [d for d in dirnames if d not in IGNORED_FOLDERS]

        for filename in filenames:
            if filename.lower().endswith(".qpsix"):
                filepath = os.path.join(dirpath, filename)

                rel_path = os.path.relpath(filepath, input_folder)
                safe_name = rel_path.replace(os.sep, "_").replace(".qpsix", "_qpsix_max.npy")
                output_path = os.path.join(output_folder, safe_name)

                if os.path.exists(output_path):
                    print(f"Already converted: {safe_name}")
                    continue

                print(f"Converting: {filepath}")
                try:
                    zdata = ZygoData(filepath)
                    img = np.std(zdata.data, axis=0)
                    np.save(output_path, img)
                    print(f"  → Saved: {output_path}")
                    any_processed = True

                except Exception as e:
                    print(f"  Error for {filepath}: {e}")

    if not any_processed:
        print("No new files to convert.")


# ================================================================
# =================== FIDUCIAL MEASUREMENT =======================
# ================================================================

def interactive_fiducial_measurement(image_path, file_name, max_num_circles=3, percentile_range=(0, 100)):
    """
    Shows an image and lets user click points to fit fiducials manually.
    SHIFT = click points
    X     = validate circle
    BACKSPACE = undo last point
    """
    try:
        image_data = np.load(image_path)
    except Exception as e:
        print(f"Couldn't load image: {e}")
        return []

    fig, ax = plt.subplots(figsize=(15, 15))

    # --- Force zoom ON at startup ---
    toolbar = plt.get_current_fig_manager().toolbar
    if toolbar and toolbar.mode != 'zoom rect':
        toolbar.zoom()

    vmin, vmax = np.percentile(image_data, percentile_range)
    ax.imshow(image_data, origin='lower', cmap='gray', vmin=vmin, vmax=vmax)
    ax.set_title(
        f"Working on: {file_name}\n"
        "SHIFT + Click = add points | X = validate circle | Backspace = undo"
    )
    plt.axis("image")

    clicked_points = []
    finished_circles = []
    point_plots = []

    preview_lines = []
    preview_circle = None

    shift_pressed = False
    zoom_was_active = False

    # ------------------------------------------------------------
    def redraw_preview():
        nonlocal preview_circle, preview_lines

        if preview_circle is not None:
            preview_circle.remove()
            preview_circle = None

        for line in preview_lines:
            line.remove()
        preview_lines = []

        if len(clicked_points) >= 3:
            try:
                x = np.array([p[0] for p in clicked_points])
                y = np.array([p[1] for p in clicked_points])
                cx, cy, r = fit_circle_to_points(x, y)

                preview_circle = Circle(
                    (cx, cy), r,
                    fill=False, color='cyan',
                    linestyle='--', linewidth=2
                )
                ax.add_patch(preview_circle)

                h, = ax.plot([cx - r, cx + r], [cy, cy], 'm--', lw=1)
                v, = ax.plot([cx, cx], [cy - r, cy + r], 'm--', lw=1)
                preview_lines = [h, v]

            except Exception:
                pass

        fig.canvas.draw_idle()

    # ------------------------------------------------------------
    def handle_mouse_click(event):
        if event.inaxes != ax or event.button != 1:
            return

        if shift_pressed:
            clicked_points.append((event.xdata, event.ydata))
            p, = ax.plot(event.xdata, event.ydata, 'bo', ms=2)
            point_plots.append(p)
            redraw_preview()

    # ------------------------------------------------------------
    def handle_keypress(event):
        nonlocal shift_pressed, zoom_was_active

        if event.key == 'shift':
            shift_pressed = True
            if toolbar and toolbar.mode == 'zoom rect':
                zoom_was_active = True
                toolbar.zoom()
            else:
                zoom_was_active = False

    # ------------------------------------------------------------
    def handle_keyrelease(event):
        nonlocal shift_pressed, zoom_was_active

        if event.key == 'shift':
            shift_pressed = False
            if zoom_was_active and toolbar:
                toolbar.zoom()
                zoom_was_active = False

        elif event.key == 'x':
            if len(clicked_points) < 3:
                print("Need at least 3 points to fit a circle")
                return

            try:
                x = np.array([p[0] for p in clicked_points])
                y = np.array([p[1] for p in clicked_points])
                cx, cy, r = fit_circle_to_points(x, y)

                idx = len(finished_circles) + 1
                print(
                    f"{file_name} | Circle #{idx}: "
                    f"center=({cx:.2f}, {cy:.2f}), radius={r:.2f}"
                )

                finished_circles.append((cx, cy))

                clicked_points.clear()
                for p in point_plots:
                    p.remove()
                point_plots.clear()

                final_circle = Circle(
                    (cx, cy), r,
                    fill=False, color='lime', linewidth=2
                )
                ax.add_patch(final_circle)

                ax.set_xlim(0, image_data.shape[1])
                ax.set_ylim(0, image_data.shape[0])
                fig.canvas.draw_idle()

                if toolbar and toolbar.mode != 'zoom rect':
                    toolbar.zoom()

                if len(finished_circles) >= max_num_circles:
                    print("All circles done. Close the window to continue.")
                    plt.close(fig)

            except Exception as e:
                print(f"Error fitting circle: {e}")

        elif event.key == 'backspace':
            if clicked_points:
                clicked_points.pop()
                if point_plots:
                    point_plots.pop().remove()
                redraw_preview()
            else:
                print("No points to undo")

    # ------------------------------------------------------------
    fig.canvas.mpl_connect('button_press_event', handle_mouse_click)
    fig.canvas.mpl_connect('key_press_event', handle_keypress)
    fig.canvas.mpl_connect('key_release_event', handle_keyrelease)

    plt.show(block=True)
    return finished_circles


def fill_template_with_fiducials(fiducials_excel_path, template_dest):
    """
    Reads fiducials_measurements.xlsx (skipping first row header and first column),
    and pastes all data rows into the 'fit' sheet of the template xlsm starting at row 34.
    """
    try:
        import openpyxl

        # Read fiducials excel, skip first column
        fid_df = pd.read_excel(fiducials_excel_path)
        data = fid_df.iloc[:, 1:].values  # skip first column

        # Open template (keep_vba=True to preserve macros)
        wb = openpyxl.load_workbook(template_dest, keep_vba=True)

        if "fit" not in wb.sheetnames:
            print(f"  Sheet 'fit' not found in template. Available sheets: {wb.sheetnames}")
            return

        ws = wb["fit"]

        START_ROW = 34
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=START_ROW + row_idx, column=1 + col_idx, value=value)

        wb.save(template_dest)
        print(f"Template filled with {len(data)} rows starting at row {START_ROW}.")

    except Exception as e:
        print(f"Could not fill template: {e}")


def save_results_to_excel(measurement_results, excel_output_path):
    """
    Saves measurement results to Excel with the correct column order:
    filename_npy | mirror | xc_1 | yc_1 | ... | xc_N | yc_N | roll | binning | filename
    """
    results_df = pd.DataFrame(measurement_results)

    yc_cols = [c for c in results_df.columns if c.startswith("yc_")]
    if yc_cols:
        last_yc = yc_cols[-1]
        insert_pos = results_df.columns.get_loc(last_yc) + 1
    else:
        insert_pos = len(results_df.columns)

    if "roll" not in results_df.columns:
        results_df.insert(insert_pos, "roll", 0)
        insert_pos += 1

    if "binning" not in results_df.columns:
        results_df.insert(insert_pos, "binning", 1)

    if "filename" in results_df.columns:
        cols = [c for c in results_df.columns if c != "filename"] + ["filename"]
        results_df = results_df[cols]

    results_df.to_excel(excel_output_path, index=False)


# ================================================================
# ========================= MAIN SCRIPT ==========================
# ================================================================

print("=== Fiducials Fitting Tool ===")

folder_path = input("️  Enter folder path containing datx/qpsix files: ").strip().strip("'\"")
npy_folder = os.path.join(folder_path, "npy_exports")

# --- Copy template xlsm if not already present ---
TEMPLATE_PATH = "/Volumes/solarc/02- Engineering/08 - Metrology/01 - Optics/07 - Measurements/FM/substrates_template_FM_form_casquette.xlsm"
template_dest = os.path.join(folder_path, os.path.basename(TEMPLATE_PATH))

if os.path.exists(template_dest):
    print(f"Template already present: {os.path.basename(TEMPLATE_PATH)}")
else:
    try:
        shutil.copy2(TEMPLATE_PATH, template_dest)
        print(f"Template copied to: {template_dest}")
    except Exception as e:
        print(f"Could not copy template: {e}")

image_type = input("Image type (tilt / sfe): ").strip().lower()
percentile_range = (5, 95) if image_type == "sfe" else (0, 100)
print(f"Using percentile range: {percentile_range[0]} - {percentile_range[1]}")

# --- Conversion step (runs only for files not yet converted) ---
print("\n=== Checking / Running conversion step ===")
convert_qpsix_to_npy(folder_path, npy_folder)

# --- Fiducial measurement ---
circles_per_image = 3
excel_output_path = os.path.join(npy_folder, "fiducials_measurements.xlsx")

if os.path.isfile(excel_output_path):
    previous_results = pd.read_excel(excel_output_path)
    processed_files = set(previous_results["filename_npy"].tolist())
    measurement_results = previous_results.to_dict(orient="records")
    print(f"Loaded existing Excel with {len(processed_files)} measured files.")
else:
    processed_files = set()
    measurement_results = []
    print("No existing Excel found, starting a new one.")

npy_file_list = [f for f in os.listdir(npy_folder) if f.lower().endswith("qpsix_max.npy")]
if not npy_file_list:
    print("No NPY files found in the npy_exports folder.")
    exit()

print(f"Found {len(npy_file_list)} NPY files ({len(processed_files)} already measured).")
print("\nHow to use:")
print(f"  • SHIFT + Click = add points to define a fiducial")
print(f"  • X = validate and save the circle (need {circles_per_image} total)")
print(f"  • Backspace = undo last click\n")

for current_file in npy_file_list:
    if current_file in processed_files:
        print(f"  Skipping already processed: {current_file}")
        continue

    print(f"\nProcessing: {current_file}")
    full_image_path = os.path.join(npy_folder, current_file)

    measured_fiducials = interactive_fiducial_measurement(
        full_image_path, current_file, circles_per_image, percentile_range=percentile_range
    )

    data_row = {
        "filename_npy": current_file,
        "mirror": current_file.split("_")[0],
    }

    for i, (cx, cy) in enumerate(measured_fiducials, start=1):
        data_row[f"xc_{i}"] = cx
        data_row[f"yc_{i}"] = cy

    data_row["filename"] = current_file.replace("_qpsix_max.npy", ".datx").replace("_", "/", 1)

    measurement_results.append(data_row)
    processed_files.add(current_file)

    save_results_to_excel(measurement_results, excel_output_path)
    fill_template_with_fiducials(excel_output_path, template_dest)

print("\nAll files processed!")